from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

BLUE  = "0C3260"
YELLOW = "FDC300"
LIGHT_BLUE = "E8EFF9"
LIGHT_YELLOW = "FFFBE6"
WHITE = "FFFFFF"
GRAY = "F4F4F4"
DARK_GRAY = "888888"

thin = Side(style='thin', color="CCCCCC")
med  = Side(style='medium', color=BLUE)
border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
border_header = Border(left=med, right=med, top=med, bottom=med)

def header_font(size=11):
    return Font(name='Arial', bold=True, color=WHITE, size=size)

def title_font():
    return Font(name='Arial', bold=True, color=BLUE, size=14)

def label_font():
    return Font(name='Arial', bold=True, color=BLUE, size=10)

def data_font(color="000000"):
    return Font(name='Arial', size=10, color=color)

def note_font():
    return Font(name='Arial', italic=True, size=9, color=DARK_GRAY)

def hfill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def left():
    return Alignment(horizontal='left', vertical='center', wrap_text=True)

# ─────────────────────────────────────────
# PLANTILLA 1: VENTAS
# ─────────────────────────────────────────
wb_v = Workbook()
ws = wb_v.active
ws.title = "Ventas Mensuales"

# Row 1: Title
ws.merge_cells("A1:G1")
ws["A1"] = "OROCash — Plantilla Importación de Ventas Mensuales"
ws["A1"].font = Font(name='Arial', bold=True, color=WHITE, size=14)
ws["A1"].fill = hfill(BLUE)
ws["A1"].alignment = center()
ws.row_dimensions[1].height = 36

# Row 2: Instructions
ws.merge_cells("A2:G2")
ws["A2"] = "Completa cada fila con los datos mensuales. Una fila = un mes. Guarda como .csv o .xlsx y súbelo al sistema."
ws["A2"].font = note_font()
ws["A2"].fill = hfill(LIGHT_BLUE)
ws["A2"].alignment = left()
ws.row_dimensions[2].height = 22

# Row 3: blank spacer
ws.row_dimensions[3].height = 6

# Row 4: Headers
headers = [
    ("Mes",               18, YELLOW, BLUE),
    ("Año",               10, YELLOW, BLUE),
    ("Ventas ($)",        16, YELLOW, BLUE),
    ("N° Joyas Vendidas", 20, YELLOW, BLUE),
    ("Ticket Promedio ($)",20, YELLOW, BLUE),
    ("Conversión (%)",    18, YELLOW, BLUE),
    ("Gramos Vendidos",   18, YELLOW, BLUE),
]
for col_idx, (label, width, bg, fg) in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col_idx, value=label)
    cell.font = Font(name='Arial', bold=True, color=fg, size=10)
    cell.fill = hfill(bg)
    cell.alignment = center()
    cell.border = border_thin
    ws.column_dimensions[get_column_letter(col_idx)].width = width
ws.row_dimensions[4].height = 30

# Months list for column A validation
meses = ["Enero","Febrero","Marzo","Abril","Mayo","Junio",
         "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]

# Data rows 5–16 (one per month as example)
sample = [
    ("Enero",   2026, None, None, None, None, None),
    ("Febrero", 2026, None, None, None, None, None),
    ("Marzo",   2026, None, None, None, None, None),
    ("Abril",   2026, None, None, None, None, None),
    ("Mayo",    2026, None, None, None, None, None),
    ("Junio",   2026, None, None, None, None, None),
    ("Julio",   2026, None, None, None, None, None),
    ("Agosto",  2026, None, None, None, None, None),
    ("Septiembre",2026,None,None, None, None, None),
    ("Octubre", 2026, None, None, None, None, None),
    ("Noviembre",2026,None, None, None, None, None),
    ("Diciembre",2026,None, None, None, None, None),
]
for r, row_data in enumerate(sample, 5):
    fill = hfill(LIGHT_YELLOW) if r % 2 == 0 else hfill(WHITE)
    for c, val in enumerate(row_data, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = data_font()
        cell.fill = fill
        cell.alignment = center()
        cell.border = border_thin
    ws.row_dimensions[r].height = 22

# Data validation: Mes dropdown
dv_mes = DataValidation(
    type="list",
    formula1='"Enero,Febrero,Marzo,Abril,Mayo,Junio,Julio,Agosto,Septiembre,Octubre,Noviembre,Diciembre"',
    allow_blank=True, showErrorMessage=True,
    error="Usa el nombre del mes en español.", errorTitle="Mes inválido"
)
ws.add_data_validation(dv_mes)
dv_mes.sqref = "A5:A200"

# Number format hints row
ws.row_dimensions[17].height = 18
ws.merge_cells("A17:G17")
ws["A17"] = "▸ Formato: Ventas y Ticket en números (ej: 28400). Conversión en decimal (ej: 2.8 para 2.8%). Gramos en enteros."
ws["A17"].font = note_font()
ws["A17"].fill = hfill(LIGHT_BLUE)
ws["A17"].alignment = left()

# Freeze header
ws.freeze_panes = "A5"

wb_v.save("/sessions/adoring-nice-faraday/mnt/outputs/plantilla_ventas_orocash.xlsx")
print("Ventas OK")

# ─────────────────────────────────────────
# PLANTILLA 2: TAREAS
# ─────────────────────────────────────────
wb_t = Workbook()
ws2 = wb_t.active
ws2.title = "Tareas"

# Row 1: Title
ws2.merge_cells("A1:G1")
ws2["A1"] = "OROCash — Plantilla Importación Masiva de Tareas"
ws2["A1"].font = Font(name='Arial', bold=True, color=WHITE, size=14)
ws2["A1"].fill = hfill(BLUE)
ws2["A1"].alignment = center()
ws2.row_dimensions[1].height = 36

# Row 2: Instructions
ws2.merge_cells("A2:G2")
ws2["A2"] = "Llena una fila por tarea. Los campos Título y Área son obligatorios. Guarda como .csv y súbelo al sistema."
ws2["A2"].font = note_font()
ws2["A2"].fill = hfill(LIGHT_BLUE)
ws2["A2"].alignment = left()
ws2.row_dimensions[2].height = 22
ws2.row_dimensions[3].height = 6

# Row 4: Headers
headers2 = [
    ("Título",             30),
    ("Descripción",        40),
    ("Área / Dept.",       22),
    ("Asignado a",         22),
    ("Prioridad",          14),
    ("Fecha límite",       16),
    ("Estado",             16),
]
for col_idx, (label, width) in enumerate(headers2, 1):
    cell = ws2.cell(row=4, column=col_idx, value=label)
    cell.font = Font(name='Arial', bold=True, color=BLUE, size=10)
    cell.fill = hfill(YELLOW)
    cell.alignment = center()
    cell.border = border_thin
    ws2.column_dimensions[get_column_letter(col_idx)].width = width
ws2.row_dimensions[4].height = 30

# Sample rows
areas = ["Fotografía","Diseño","Prod. Audiovisual","Social Media","Ventas Online",
         "Ventas Mariela","Publicidad Móvil","Marketing","Eventos / Vitrinimos","General"]
personas = ["Israel Pacheco","Arturo García","Gustavo Palma","Gissella Barzola",
            "Pamela Mosquera","Mariela García","Katherine Escobar","Eddy Mendoza",
            "Angel Molineros","Miguel Montoya","Jaqueline Jara","Omairo Bueno","David Choez"]
prioridades = ["alta","media","baja"]
estados = ["pendiente","en_progreso","completado","vencida"]

samples_t = [
    ("Sesión fotográfica nueva línea","Fotos de nueva colección verano","Fotografía","Israel Pacheco","alta","2026-07-10","pendiente"),
    ("Diseño catálogo julio","Catálogo digital segunda mitad 2026","Diseño","Gissella Barzola","media","2026-07-05","pendiente"),
    ("Video institucional Q3","Producción video para redes","Prod. Audiovisual","Arturo García","media","2026-07-08","en_progreso"),
    ("Pautaje agosto","Planificación pauta digital agosto","Social Media","Pamela Mosquera","alta","2026-07-28","pendiente"),
]
for r, row_data in enumerate(samples_t, 5):
    fill = hfill(GRAY) if r % 2 == 0 else hfill(WHITE)
    for c, val in enumerate(row_data, 1):
        cell = ws2.cell(row=r, column=c, value=val)
        cell.font = data_font()
        cell.fill = fill
        cell.alignment = left() if c <= 2 else center()
        cell.border = border_thin
    ws2.row_dimensions[r].height = 22

# Empty rows for user to fill
for r in range(9, 110):
    fill = hfill(GRAY) if r % 2 == 0 else hfill(WHITE)
    for c in range(1, 8):
        cell = ws2.cell(row=r, column=c, value=None)
        cell.fill = fill
        cell.border = border_thin
        cell.alignment = center()
        cell.font = data_font()
    ws2.row_dimensions[r].height = 20

# Data validations
dv_area = DataValidation(type="list",
    formula1='"Fotografía,Diseño,Prod. Audiovisual,Social Media,Ventas Online,Ventas Mariela,Publicidad Móvil,Marketing,Eventos / Vitrinimos,General"',
    allow_blank=True, showErrorMessage=True,
    error="Usa una de las áreas de la lista.", errorTitle="Área inválida")
ws2.add_data_validation(dv_area)
dv_area.sqref = "C5:C200"

dv_prio = DataValidation(type="list", formula1='"alta,media,baja"',
    allow_blank=True, showErrorMessage=True,
    error="Usa: alta, media o baja.", errorTitle="Prioridad inválida")
ws2.add_data_validation(dv_prio)
dv_prio.sqref = "E5:E200"

dv_stat = DataValidation(type="list",
    formula1='"pendiente,en_progreso,completado,vencida"',
    allow_blank=True, showErrorMessage=True,
    error="Usa uno de los estados disponibles.", errorTitle="Estado inválido")
ws2.add_data_validation(dv_stat)
dv_stat.sqref = "G5:G200"

# Note row
ws2.row_dimensions[110].height = 18
ws2.merge_cells("A110:G110")
ws2["A110"] = "▸ Asignado a: usa el nombre completo exacto de la persona. Fecha límite: formato YYYY-MM-DD (ej: 2026-07-15)."
ws2["A110"].font = note_font()
ws2["A110"].fill = hfill(LIGHT_BLUE)
ws2["A110"].alignment = left()

# Reference sheet
ws_ref = wb_t.create_sheet("Referencia")
ws_ref["A1"] = "Personas del equipo"
ws_ref["A1"].font = Font(name='Arial', bold=True, color=WHITE, size=11)
ws_ref["A1"].fill = hfill(BLUE)
ws_ref["A1"].alignment = center()
ws_ref.column_dimensions["A"].width = 25
ws_ref.column_dimensions["B"].width = 30
ws_ref.column_dimensions["C"].width = 25

ws_ref["B1"] = "Áreas disponibles"
ws_ref["B1"].font = Font(name='Arial', bold=True, color=WHITE, size=11)
ws_ref["B1"].fill = hfill(BLUE)
ws_ref["B1"].alignment = center()

ws_ref["C1"] = "Estados"
ws_ref["C1"].font = Font(name='Arial', bold=True, color=WHITE, size=11)
ws_ref["C1"].fill = hfill(BLUE)
ws_ref["C1"].alignment = center()
ws_ref.row_dimensions[1].height = 28

for i, p in enumerate(personas, 2):
    c = ws_ref.cell(row=i, column=1, value=p)
    c.font = data_font()
    c.fill = hfill(LIGHT_BLUE) if i%2==0 else hfill(WHITE)
    c.alignment = left()
for i, a in enumerate(areas, 2):
    c = ws_ref.cell(row=i, column=2, value=a)
    c.font = data_font()
    c.fill = hfill(LIGHT_YELLOW) if i%2==0 else hfill(WHITE)
    c.alignment = left()
for i, s in enumerate(["pendiente","en_progreso","completado","vencida","","alta","media","baja"], 2):
    c = ws_ref.cell(row=i, column=3, value=s)
    c.font = data_font()
    c.fill = hfill(GRAY) if i%2==0 else hfill(WHITE)
    c.alignment = left()

ws2.freeze_panes = "A5"

wb_t.save("/sessions/adoring-nice-faraday/mnt/outputs/plantilla_tareas_orocash.xlsx")
print("Tareas OK")
